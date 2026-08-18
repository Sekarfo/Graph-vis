"""Факт СМР из «общий свод по смр СНП 2.0.xlsx» → smr2/<область>.json.

МОДЕЛЬ МАРШРУТА. Свод описывает не только длину, но и сам маршрут: столбцы
10..21 идут парами (метка, км) и читаются как цепочка

    L1 →k1→ L2 →k2→ ... →Ln→kn→ НП

где метка — объект, ОТ которого идёт отрезок, а последний отрезок упирается
в сам населённый пункт. Проверено на Жамбылской области: сумма отрезков
сходится с «Общая протяженность до НП» во всех 169 строках из 169, а путь по
графу от первой метки до НП совпадает с планом (±0.35 км) в 95 строках из 128
разрешимых. Поэтому цепочка — ОСНОВНОЙ источник маршрута, а «Подключение к
сети СПД» (ПкСС) используется лишь как подсказка направления и запасной якорь.

Отсюда два штатных расхождения между планом свода и графом, оба описаны в
самом своде и оба обрабатываются здесь:

  * граф ДЛИННЕЕ плана — маршрут ПкСС→НП проходит через другие сёла, и часть
    км уже отнесли к ним. Нам принадлежит только ХВОСТ длиной в план,
    упирающийся в НП (см. tail_of).
  * граф КОРОЧЕ плана — работы вели и к самому ПкСС, поэтому цепочка
    начинается ВЫШЕ него по сети (в Жамбыле это стр. 10: П1 →15→ РРС Жума
    →9→ П2 →2→ Құмсуат, где РРС Жума и есть ПкСС). Излишек продлевается
    вверх по сети, пока путь однозначен (см. extend_upstream); на развилке
    ничего не красим и пишем строку в отчёт для ручного разбора.

ФАКТ — ровно два столбца-итога: [614] «Магистр. Сеть (трубка) км» и
[805] «Магистр. Сеть (ВОЛС) км» (накопительный итог; помесячные и посуточные
колонки внутри них — слагаемые, отдельно не читаются). Сначала прокладывают
трубку, затем задувают в неё ВОЛС, поэтому на фронте это два независимых
режима показа, а не одна метрика (см. дропбокс «Показывать прогресс»).

Запуск:  python scripts/import_smr2.py [--region zhambyl] [--report-only]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

from matcher import normalize_name          # noqa: E402
from progress_core import (                  # noqa: E402
    HEAD_KINDS, PLANNED_TYPES, Graph, fill_along,
)

REGIONS_DIR = BASE / "regions"
OUT_DIR = BASE / "smr2"
REPORT = BASE / "SMR2_RECONCILE.md"

# --- раскладка листа ------------------------------------------------------
# Номера столбцов НЕ зашиты: книгу правят руками, и столбцы в ней уже уезжали.
# Между локальной копией от 14.08 и текущей таблицей в Google вставили три
# столбца («ДХ/ГУ по запросу»), из-за чего итог трубки переехал с WQ на WT —
# по старым индексам мы прочитали бы соседнюю пустую колонку и молча потеряли
# факт (так по с.Терсакан выходило 14.242 вместо 30.123). Поэтому каждый
# столбец ищется по заголовку, см. locate_columns().
HDR_REGION, HDR_DISTRICT = "Область", "Район"
HDR_OKRUG, HDR_NAME = "Сельский округ", "Нас пункт"
HDR_HEAD, HDR_PLAN, HDR_TECH = "Подключение к сети СПД", "Общая протяженность до НП", "Технология"
HDR_SEG_KM = "протяженность км"          # повторяется у каждого участка
HDR_TUBE, HDR_FIBER = "Магистр. Сеть (трубка) км", "Магистр. Сеть (ВОЛС) км"

# Ссылка на живую таблицу: экспорт открыт по ссылке, авторизация не нужна.
SHEET_ID = "1hI6neYEmaB8x5vJsNfEXaDtJ_x5ToS_jt9Ja1T4EcZU"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"


def _norm_hdr(v) -> str:
    """Заголовок к сравнимому виду: в книге они с переносами строк и хвостовыми
    пробелами («Утвержд. акт выбора трассы »)."""
    return " ".join(str(v).split()).strip().lower() if v is not None else ""


class Columns:
    """Индексы столбцов конкретного листа, найденные по заголовкам."""

    def __init__(self, header: tuple):
        idx: dict[str, int] = {}
        seg_km: list[int] = []
        for c, v in enumerate(header):
            h = _norm_hdr(v)
            if not h:
                continue
            if h == _norm_hdr(HDR_SEG_KM):
                seg_km.append(c)
            elif h not in idx:            # берём ПЕРВОЕ вхождение: «Распред.
                idx[h] = c                # Сеть (ВОЛС) км» в книге задвоена
        def need(title: str) -> int:
            c = idx.get(_norm_hdr(title))
            if c is None:
                raise KeyError(f"в шапке листа нет столбца «{title}»")
            return c
        self.region = need(HDR_REGION)
        # Блок опознавания села (№ | Область | Район | Сельский округ | Нас пункт)
        # идёт подряд во всех листах книги, но подписи местами стёрты — в
        # Костанайской пустует заголовок над столбцом района, хотя данные там
        # есть. Отсчитываем от «Области», раз уж она нашлась.
        self.district = idx.get(_norm_hdr(HDR_DISTRICT), self.region + 1)
        self.okrug = idx.get(_norm_hdr(HDR_OKRUG), self.region + 2)
        self.name = need(HDR_NAME)
        self.head = need(HDR_HEAD)
        self.plan = need(HDR_PLAN)
        self.tech = need(HDR_TECH)
        self.tube = need(HDR_TUBE)
        self.fiber = need(HDR_FIBER)
        # Метка участка стоит в столбце слева от его километража.
        self.segs = [(c - 1, c) for c in seg_km]
        if not self.segs:
            raise KeyError(f"в шапке листа нет ни одного столбца «{HDR_SEG_KM}»")

SHEET_TO_SLUG = {
    "Туркестанская область": "turkestan",
    "Жамбылская область": "zhambyl",
    "Кызылординская область": "kyzylorda",
    "Алматинская область": "almaty-zhetysu",
    "область Жетису": "almaty-zhetysu",
    "ВКО": "vko",
    "Павлодарская область": "pavlodar",
    "область Абай": "abai",
    "Мангистауская область": "mangystau-atyrau",
    "Актюбинская область": "aktobe",
    "Атырауская область": "mangystau-atyrau",
    "Западно-Казахстанская область": "zko",
    "Костанайская область": "kostanay",
    "Акмолинская область": "akmola",
    "Северо-Казахстанская область": "sko",
    "Карагандинская область": "karaganda-ulytau",
    "Область Ұлытау": "karaganda-ulytau",
}

# Допуск на «длина сошлась»: в своде км округлены до 0.1, в чертеже — тоже,
# поэтому на цепочке из нескольких отрезков накапливается до ~0.3 км.
TOL = 0.35
# Порог привязки метки к узлу графа. Ниже — считаем метку неопознанной и не
# строим по ней маршрут: лучше отправить строку в отчёт, чем закрасить чужое
# ребро (тот же BIND_MIN, что и у привязки отчётов бота).
LABEL_MIN = 80

# «П1», «П-1», «П 1» — безымянная муфта: в чертеже у неё нет своего имени,
# в графе ей соответствует узел, который мы находим ходом по маршруту, а не
# по названию. А вот «М1», «М18/1» — это НАСТОЯЩИЕ имена муфт графа.
_PLACEHOLDER_RE = re.compile(r"^\s*п\s*-?\s*\d+\s*$", re.IGNORECASE)


def is_placeholder(label: str | None) -> bool:
    return bool(label) and bool(_PLACEHOLDER_RE.match(str(label)))


def num(v) -> float | None:
    """Число из ячейки; свод местами хранит км текстом («12,5», «12 »)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", ".").replace(" ", "").replace("\xa0", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def clean(v) -> str:
    return "" if v is None else str(v).strip()


# ---------------------------------------------------------------- разбор ---
def parse_sheet(ws) -> tuple[list[dict], dict]:
    """Строки листа области. Шапка занимает 2 строки, ниже данные, а дальше
    итоговая строка — и под ней ВТОРАЯ таблица («ВОЛС в рамках СНП 38ТП»,
    пункты пропуска) с другой раскладкой столбцов. Её нельзя принимать за сёла,
    поэтому читаем строго до итогов и обрываемся.

    Итоговая строка не подписана словом «Итого»: в ней вместо названия области
    стоит счётчик (Жамбыл: [164, 1, 10, 89, 164, 55, 1412.5, 84]).
    Отсюда признак конца — «Область» перестала быть текстом."""
    recs: list[dict] = []
    control: dict = {}
    cols: Columns | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            cols = Columns(row)
            continue
        if i == 2:
            continue

        def cell(c: int):
            return row[c] if c < len(row) else None

        region = clean(cell(cols.region))
        name = clean(cell(cols.name))
        if name.lower().startswith("итого") or region.lower().startswith("итого")                 or (region and num(cell(cols.region)) is not None):
            # Итоги листа — независимая проверка нашего разбора (см. main).
            control = {
                "row": i,
                "np": num(cell(cols.name)),
                "planKm": num(cell(cols.plan)),
                "tubeKm": num(cell(cols.tube)),
                "fiberKm": num(cell(cols.fiber)),
            }
            break                        # итоговая строка → таблица кончилась
        if not name or not region:
            continue

        segs = []
        for lc, kc in cols.segs:
            label, km = clean(cell(lc)), num(cell(kc))
            if not label and not km:
                continue
            segs.append({"label": label, "km": km or 0.0})

        recs.append({
            "row": i,
            "name": name,
            "district": clean(cell(cols.district)),
            "okrug": clean(cell(cols.okrug)),
            "head": clean(cell(cols.head)),
            "tech": clean(cell(cols.tech)),
            "planKm": num(cell(cols.plan)) or 0.0,
            "segs": segs,
            "tubeKm": num(cell(cols.tube)) or 0.0,
            "fiberKm": num(cell(cols.fiber)) or 0.0,
        })
    return recs, control


# ------------------------------------------------------------- маршруты ---
def other_end(edge: dict, node_id: str) -> str:
    return edge["to"] if edge["from"] == node_id else edge["from"]


def components(graph: Graph) -> dict[str, int]:
    """Номер компоненты связности для каждого узла (считается один раз на граф).

    Узлы из разных компонент маршрутом не соединить, поэтому кандидат из чужой
    компоненты — всегда промах матчера, каким бы похожим ни было название."""
    cached = getattr(graph, "_smr2_components", None)
    if cached is not None:
        return cached
    comp: dict[str, int] = {}
    idx = 0
    for start in graph.nodes:
        if start in comp:
            continue
        idx += 1
        stack = [start]
        while stack:
            u = stack.pop()
            if u in comp:
                continue
            comp[u] = idx
            for e in graph.adj.get(u, ()):
                v = other_end(e, u)
                if v not in comp:
                    stack.append(v)
    graph._smr2_components = comp
    return comp


def node_label(graph: Graph, node_id: str) -> str:
    """Человекочитаемое имя узла для отчёта: часть муфт в чертеже безымянна,
    и «None» в тексте разбирать невозможно — показываем kind и id."""
    n = graph.nodes.get(node_id) or {}
    return n.get("name") or f"{n.get('kind', 'узел')} {node_id}"


def node_seq(path: list[dict], start_id: str) -> list[str]:
    """Последовательность узлов вдоль упорядоченного пути (len = len(path)+1)."""
    seq, cur = [start_id], start_id
    for e in path:
        cur = other_end(e, cur)
        seq.append(cur)
    return seq


def path_km(path: list[dict]) -> float:
    return sum(e.get("lengthKm") or 0.0 for e in path)


def tail_of(path: list[dict], start_id: str, km: float) -> tuple[list[dict], str, bool]:
    """Хвост маршрута длиной ~km, упирающийся в КОНЕЦ пути (в НП).

    Возвращает (рёбра хвоста по ходу к НП, узел-начало хвоста, легла ли
    граница ровно на узел). Граница, не легшая на узел, — сигнал, что план
    свода режет ребро посередине: закрасить можно, но в отчёт это идёт."""
    seq = node_seq(path, start_id)
    acc, take = 0.0, 0
    for e in reversed(path):
        acc += e.get("lengthKm") or 0.0
        take += 1
        if acc >= km - TOL:
            break
    exact = abs(acc - km) <= TOL
    cut = len(path) - take
    return path[cut:], seq[cut], exact


def upstream_guide(segs: list[dict], consumed_km: float) -> list[float]:
    """Длины участков свода в порядке движения ВВЕРХ от НП, за вычетом того,
    что уже покрыл найденный коридор. Это подсказка для выбора на развилках:
    свод знает, каким по длине должен быть следующий отрезок."""
    out: list[float] = []
    left = consumed_km
    for km in (s["km"] for s in reversed(segs) if s["km"]):
        if left > TOL:
            if km <= left + TOL:
                left -= km
                continue
            km -= left
            left = 0.0
        out.append(km)
    return out


def extend_upstream(graph: Graph, anchor_id: str, first_edge_id: str | None,
                    km: float, guide: list[float] | None = None
                    ) -> tuple[list[dict] | None, str]:
    """Продлить коридор ВВЕРХ от якоря на km (случай «граф короче плана»).

    Идём от якоря в сторону, противоположную НП. На развилке решаем по двум
    правилам, в этом порядке:

      1. Длина ребра совпадает с длиной очередного участка свода. Свод задаёт
         маршрут по отрезкам, поэтому это самый прямой признак: у Ұланбела от
         муфты N1155 идут ветки 55 и 88 км, а в своде записано ровно 88 —
         значит трасса уходит на N1154, а не на N1153.
      2. Ребро плановое. По плану подключения то, что ещё не построено,
         достроить обязаны, поэтому трасса идёт по плановым линиям, а не по
         уже существующим.

    Если и это не развело варианты — возвращаем None с причиной: такие случаи
    по договорённости не красим, а выписываем в отчёт для ручного разбора."""
    path: list[dict] = []
    cur, remaining = anchor_id, km
    left = list(guide or [])
    visited = {anchor_id}
    used = {first_edge_id} if first_edge_id else set()
    # Первый шаг делаем и когда весь остаток меньше допуска: план бывает
    # короче TOL (Жанатасашы — 0.3 км), и тогда ход не начинался вовсе,
    # хотя нужное ребро в графе есть и совпадает по длине.
    while remaining > TOL or (not path and remaining > 0.01):
        opts = [e for e in graph.adj.get(cur, ())
                if e["id"] not in used and other_end(e, cur) not in visited]
        # OLT ставят спутником рядом с ATN/АТС того же названия, и в графе это
        # отдельный узел-близнец. Ребро к близнецу — не ответвление трассы, а
        # та же площадка, иначе каждый такой узел выглядел бы развилкой.
        if len(opts) > 1:
            cur_nn = normalize_name(graph.nodes[cur].get("name"))
            trunk = [e for e in opts
                     if not cur_nn
                     or normalize_name(graph.nodes[other_end(e, cur)].get("name")) != cur_nn]
            if trunk:
                opts = trunk
        if len(opts) > 1:
            target = left[0] if left else remaining
            exact = [e for e in opts if abs((e.get("lengthKm") or 0.0) - target) <= TOL]
            if len(exact) == 1:
                opts = exact
        if len(opts) > 1:
            planned = [e for e in opts if e.get("type") in PLANNED_TYPES]
            if len(planned) == 1:
                opts = planned
        if not opts:
            return None, "тупик: выше по сети продолжения нет"
        if len(opts) > 1:
            names = ", ".join(f"«{node_label(graph, other_end(e, cur))}» "
                              f"({e.get('lengthKm')} км)" for e in opts[:4])
            return None, f"развилка из {len(opts)}: {names}"
        e = opts[0]
        path.append(e)
        used.add(e["id"])
        cur = other_end(e, cur)
        visited.add(cur)
        length = e.get("lengthKm") or 0.0
        remaining -= length
        while length > TOL and left:      # участок свода мог лечь на два ребра
            take = min(length, left[0])
            left[0] -= take
            length -= take
            if left[0] <= TOL:
                left.pop(0)
    path.reverse()               # по ходу движения: от дальнего конца к якорю
    return path, ""


# ------------------------------------------------------------ привязка ---
# Свод часто подписывает село как «НовоеИмя (СтароеИмя)», а в графе стоит только
# одно из двух — причём иногда слитно («с.Жайлау кол (Победа)» против узла
# «Жайлаукол»). Поэтому пробуем строку целиком, без скобки и только скобку.
_PAREN_RE = re.compile(r"^([^()]+)\((.+?)\)\s*$")


def name_variants(raw: str) -> list[str]:
    out = [raw.strip()]
    m = _PAREN_RE.match(raw.strip())
    if m:
        out += [m.group(1).strip(), m.group(2).strip()]
    return [v for v in out if v]


def name_index(graph: Graph) -> list[tuple[str, str, dict]]:
    """(норм. имя, норм. район, узел) — считается один раз на граф.

    Нормализация запроса и узлов — самое горячее место разбора: без индекса
    node_score перенормализовал бы название узла на каждое сравнение."""
    cached = getattr(graph, "_smr2_index", None)
    if cached is not None:
        return cached
    idx = []
    for n in graph.nodes.values():
        nn = normalize_name(n.get("name"))
        if nn:
            idx.append((nn, normalize_name(n.get("district")), n))
    graph._smr2_index = idx
    return idx


def best_nodes(graph: Graph, raw: str, district: str, *,
               kinds: tuple | None = None, in_component_of: str | None = None
               ) -> tuple[float, list[dict]]:
    """(лучший счёт, все узлы с этим счётом) по названию из свода.

    Варианты названия проверяются по очереди, и каждый следующий получает
    крошечный штраф — чтобы при РАВНОМ счёте побеждала исходная строка. Иначе
    «с.Орта тобе (Береке)», где в графе узел назван так же со скобкой, уехал бы
    на постороннее «Береке» из соседнего района с тем же счётом."""
    queries = [(normalize_name(v), i * 0.001) for i, v in enumerate(name_variants(raw))]
    queries = [(q, pen) for q, pen in queries if len(q) >= 2]
    if not queries:
        return 0.0, []
    dn = normalize_name(district)
    comp = components(graph) if in_component_of else None
    best_sc, cands = 0.0, []
    for nn, ndn, n in name_index(graph):
        if kinds and n.get("kind") not in kinds:
            continue
        if comp is not None and comp.get(n["id"]) != comp.get(in_component_of):
            continue
        sc = 0.0
        for q, pen in queries:
            # graph._score — та же шкала, что у find_node_scored (бонус за snp,
            # бонус/штраф за район); берём её, чтобы привязка здесь и в
            # остальном вьюере оценивалась одинаково.
            v = graph._score(q, dn, nn, ndn, n) - pen
            if v > sc:
                sc = v
        if sc > best_sc + 1e-9:
            best_sc, cands = sc, [n]
        elif abs(sc - best_sc) <= 1e-9 and sc > 0:
            cands.append(n)
    return best_sc, cands


def nearest_of(graph: Graph, cands: list[dict], near_id: str | None) -> dict | None:
    """Из равных по названию — ближайший ПО ГРАФУ к якорю: муфты нумеруются в
    пределах ветки («М 3» встречается десятками), и участок связывает близкие
    точки, а не тёзок с другого конца области."""
    if not cands:
        return None
    if len(cands) == 1 or not near_id:
        return cands[0]

    def km_to_near(n: dict) -> float:
        path = graph.shortest_path(n["id"], near_id)
        return path_km(path) if path is not None else float("inf")

    return min(cands, key=km_to_near)


def resolve_label(graph: Graph, label: str, district: str,
                  near_id: str | None) -> dict | None:
    """Узел графа для метки участка.

    Если near_id задан, кандидат обязан быть с ним в одной компоненте связности:
    иначе матчер выбирает тёзку, до которого маршрута физически нет. В Жамбыле
    так терялся Ақжол — метку «Тасөткел» забирал изолированный узел «Тасөткел
    рзд.» из нужного района (бонус за район +8), обгоняя настоящий «Тасөткел»
    соседнего района (штраф −15), хотя у него нет ни одного ребра."""
    if not label or is_placeholder(label):
        return None
    sc, cands = best_nodes(graph, label, district)
    node = nearest_of(graph, cands, near_id) if sc >= LABEL_MIN else None
    if node is None or near_id is None:
        return node
    comp = components(graph)
    if comp.get(node["id"]) == comp.get(near_id):
        return node
    sc2, cands2 = best_nodes(graph, label, district, in_component_of=near_id)
    if sc2 >= LABEL_MIN and cands2:
        return nearest_of(graph, cands2, near_id)
    return node


def resolve_head(graph: Graph, label: str, district: str, snp: dict) -> dict | None:
    """Узел ПкСС («Подключение к сети СПД»).

    Столбец называет ИСТОЧНИК сигнала, но подпись часто совпадает с названием
    села («АТС Қызыларық»), а общий матчер даёт kind=snp бонус +6 — и головой
    оказывается само село, отчего маршрут вырождается в ноль. Причём «АТС X»
    в своде нередко означает не АТС, а OLT X. Поэтому, если голова совпала с
    селом, ищем заново среди узлов-источников: рядом с селом почти всегда есть
    одноимённая АТС/OLT. Жуалы: OLT «Кызыларык» →1→ муфта →1→ СНП «Қызыларық»,
    ровно 2 км плана свода."""
    node = resolve_label(graph, label, district, near_id=snp["id"])
    if node is not None and node["id"] != snp["id"]:
        return node
    sc, cands = best_nodes(graph, label, district, kinds=HEAD_KINDS,
                           in_component_of=snp["id"])
    if sc < LABEL_MIN:
        return node
    best, best_km = None, None
    for n in cands:
        if n["id"] == snp["id"]:
            continue
        path = graph.shortest_path(n["id"], snp["id"])
        if path is None:
            continue
        km = path_km(path)
        if best_km is None or km < best_km:
            best, best_km = n, km
    return best or node


def donor_segment(graph: Graph, recs: list[dict], rec: dict, snp: dict
                  ) -> tuple[list[dict] | None, str | None, dict | None, dict | None]:
    """Участок до села, у которого в своде нет собственной протяжённости.

    Свод перечисляет объекты вдоль маршрута, поэтому такое село почти всегда
    встречается ПРОМЕЖУТОЧНОЙ меткой в строке соседнего НП. Участок, ведущий
    к нему, — тот, что стоит в цепочке ПЕРЕД ним.

    Возвращает (рёбра коридора, дальний узел, участок-донор, строка-донор)."""
    target = normalize_name(rec["name"])
    if not target:
        return None, None, None, None
    for other in recs:
        if other is rec:
            continue
        segs = other["segs"]
        for j in range(len(segs) - 1):
            if normalize_name(segs[j + 1]["label"]) != target:
                continue
            km = segs[j]["km"]
            if not km:
                continue
            label = segs[j]["label"]
            if is_placeholder(label) or not label:
                # Донор начинается с безымянной муфты — её в графе по названию
                # не найти, поэтому просто отходим от села на нужные км.
                path, _why = extend_upstream(graph, snp["id"], None, km, [km])
                if path:
                    return path, _far_of(graph, path, snp["id"]), segs[j], other
                continue
            src = resolve_label(graph, label, other["district"], near_id=snp["id"])
            if src is None or src["id"] == snp["id"]:
                continue
            path = graph.shortest_path(src["id"], snp["id"])
            if path is None or abs(path_km(path) - km) > TOL:
                continue
            return path, src["id"], segs[j], other
    return None, None, None, None


def snp_candidates(graph: Graph, rec: dict) -> list[dict]:
    """Узлы-кандидаты для НП, лучшие по названию — вперёд.

    Тёзок среди сёл много (в Алматинской и Жетісу — 18 групп на 42 узла), и
    район их не разводит: оба «Сауыншы» стоят в Карасайском. Кто из них нужен,
    решает не название, а МАРШРУТ свода: правильный «Сауыншы» — тот, что в
    1.3 км от АТС Бекболат, как и записано в строке. Поэтому здесь только
    собираем кандидатов, а выбирает bind() — по тому, чей коридор сошёлся."""
    sc, cands = best_nodes(graph, rec["name"], rec["district"])
    return cands[:6] if sc >= LABEL_MIN else []


# Насколько привязка удачна: меньше — лучше. Точная цепочка бьёт всё, разбор
# вручную — хуже всего; внутри одного ранга выигрывает та, где длина коридора
# ближе к плану свода.
_RANK = {
    "chain": 0, "tail": 0,
    "donor_segment": 1, "tail_short": 1, "head_to_np": 1,
    "from_np": 2,
    "chain_len_mismatch": 3, "tail_inexact": 3,
    "implausible_corridor": 4, "no_route": 4, "head_not_found": 4,
    "upstream_ambiguous": 4, "skipped_no_plan": 4, "np_not_found": 5,
}


def bind(graph: Graph, rec: dict, recs: list[dict]) -> dict:
    """Привязка строки свода. При тёзках пробуем каждого и берём того, чей
    маршрут лучше сходится со сводом."""
    cands = snp_candidates(graph, rec)
    if not cands:
        return bind_to(graph, rec, None, recs)
    best, best_q = None, None
    for snp in cands:
        res = bind_to(graph, rec, snp, recs)
        q = (_RANK.get(res["status"], 5), abs(res.get("diffKm") or 0.0))
        if best_q is None or q < best_q:
            best, best_q = res, q
        if q[0] == 0 and q[1] <= TOL:
            break                      # точное совпадение — дальше не ищем
    if len(cands) > 1:
        best["homonyms"] = len(cands)
    return best


def bind_to(graph: Graph, rec: dict, snp: dict | None, recs: list[dict]) -> dict:
    """Разложить строку свода на коридор рёбер графа + диагностика."""
    res = {
        **{k: rec[k] for k in ("row", "name", "district", "okrug", "head",
                               "tech", "planKm", "tubeKm", "fiberKm")},
        "segs": rec["segs"],
        "status": "", "note": "",
        "nodeId": None, "headNodeId": None, "anchorLabel": None, "anchorNodeId": None,
        "corridorKm": 0.0, "diffKm": None, "edgeIds": [], "upstreamKm": 0.0,
        "segChecks": [], "clippedTubeKm": 0.0, "clippedFiberKm": 0.0,
        "paintedTubeKm": 0.0, "paintedFiberKm": 0.0,
        "capKm": rec["planKm"], "homonyms": 1,
    }

    if snp is None:
        res["status"] = "np_not_found"
        res["note"] = "села нет в графе (или название расходится)"
        return res
    res["nodeId"] = snp["id"]

    head = resolve_head(graph, rec["head"], rec["district"], snp)
    res["headNodeId"] = head["id"] if head else None

    plan = rec["planKm"]
    if plan <= 0:
        # Своей протяжённости у строки нет, но участок до неё обычно описан в
        # ЧУЖОЙ строке: свод перечисляет объекты по маршруту, и наше село может
        # стоять там промежуточной меткой. Алматинская, с.Копа: у с.Еспе записано
        # «НРП Таргал →33.5→ Копа →14.6→ П1 →6.2→ Еспе», значит участок до Копы —
        # те самые 33.5 км, на них и кладём её 33.2 км факта.
        corridor, far_id, seg, donor = donor_segment(graph, recs, rec, snp)
        if corridor:
            res["status"] = "donor_segment"
            res["capKm"] = seg["km"]
            res["corridorKm"] = round(path_km(corridor), 3)
            res["edgeIds"] = [e["id"] for e in corridor]
            res["note"] = (f"плана в своде нет; участок взят из строки "
                           f"{donor['row']} «{donor['name']}»: "
                           f"«{seg['label'] or 'П'}» →{seg['km']}→ «{rec['name']}»")
            res["_corridor"], res["_farId"] = corridor, far_id
            return res
        # Участка в чужих строках тоже нет — тогда просто берём отрезок между
        # ПкСС и селом и кладём факт на него. Что лежит по пути (муфты, другие
        # объекты) и какой он длины — неважно: задача найти нужный участок между
        # двумя точками. Акмолинская, с.Маншук Маметова: метка участка в своде
        # склеена в мусор «(АТС Жалгызкудук) 0 (АТС Шалкар)», но ПкСС назван —
        # «с.Шалкар», и путь Шалкар →0.4→ муфта →11.9→ Маншук в графе есть.
        if head is not None:
            path = graph.shortest_path(head["id"], snp["id"])
            if path:
                res["status"] = "head_to_np"
                res["capKm"] = path_km(path)
                res["corridorKm"] = round(res["capKm"], 3)
                res["edgeIds"] = [e["id"] for e in path]
                res["note"] = (f"плана в своде нет; взят отрезок ПкСС→село: "
                               f"«{head.get('name')}» →{res['corridorKm']}→ "
                               f"«{snp.get('name')}»")
                res["_corridor"], res["_farId"] = path, head["id"]
                return res
        res["status"] = "skipped_no_plan"
        res["note"] = ("в своде нет протяжённости до НП, участка в чужих строках "
                       "нет, а до ПкСС маршрута не нашлось — строка пропущена")
        return res

    # --- 1. Якорь: первая ИМЕНОВАННАЯ метка цепочки, опознанная в графе.
    anchor = anchor_idx = None
    for j, s in enumerate(rec["segs"]):
        n = resolve_label(graph, s["label"], rec["district"], near_id=snp["id"])
        if n is not None and n["id"] != snp["id"]:
            anchor, anchor_idx = n, j
            break

    corridor: list[dict] | None = None
    far_id = None
    guide: list[float] = []
    if anchor is not None:
        # Коридор = путь от якоря до села; ожидаемая длина — сумма отрезков
        # ОТ якоря и правее (то, что левее, лежит выше по сети — случай B).
        corridor = graph.shortest_path(anchor["id"], snp["id"])
        if corridor is None:
            res["status"] = "no_route"
            res["note"] = (f"нет пути «{anchor.get('name')}» → «{snp.get('name')}» "
                           f"(граф разорван)")
        else:
            far_id = anchor["id"]
            res["anchorLabel"] = rec["segs"][anchor_idx]["label"]
            res["anchorNodeId"] = anchor["id"]
            res["upstreamKm"] = round(sum(s["km"] for s in rec["segs"][:anchor_idx]), 3)
            guide = upstream_guide(rec["segs"][:anchor_idx], 0.0)
            expected = sum(s["km"] for s in rec["segs"][anchor_idx:])
            res["status"] = "chain" if abs(path_km(corridor) - expected) <= TOL \
                else "chain_len_mismatch"
            if res["status"] == "chain_len_mismatch":
                res["note"] = (f"путь от «{res['anchorLabel']}» до села — "
                               f"{path_km(corridor):.1f} км, свод даёт {expected:.1f} км")
    elif head is None:
        # --- 2. Именованных якорей нет (вся цепочка — П-N) и голова не опознана.
        res["status"] = "head_not_found"
        res["note"] = f"ПкСС «{rec['head']}» не опознан, якорей в цепочке нет"
    else:
        # Берём ХВОСТ пути ПкСС→НП длиной в план: всё, что раньше, свод уже
        # отнёс к другим сёлам.
        full = graph.shortest_path(head["id"], snp["id"])
        if full is None:
            res["status"] = "no_route"
            res["note"] = (f"нет пути «{head.get('name')}» → «{snp.get('name')}» "
                           f"(граф разорван)")
        else:
            total = path_km(full)
            if total + TOL < plan:
                # Путь короче плана — остаток лежит выше ПкСС (случай B).
                # Вырожденный случай: ПкСС — это сам НП либо его OLT-спутник в
                # нуле километров (Ұланбел подключают к собственному OLT). Тогда
                # идти «выше головы» некуда — OLT висит листом, — и весь маршрут
                # лежит выше САМОГО СЕЛА: якорем берём его. Магистраль к Ұланбелу
                # так и собирается: N1154 →88→ N1155 →5→ село, ровно 93 км плана.
                if total <= TOL:
                    corridor, far_id = [], snp["id"]
                    res["upstreamKm"] = round(plan, 3)
                    guide = upstream_guide(rec["segs"], 0.0)
                else:
                    corridor, far_id = full, head["id"]
                    res["upstreamKm"] = round(plan - total, 3)
                    guide = upstream_guide(rec["segs"], total)
                res["status"] = "tail_short"
                res["note"] = (f"путь ПкСС→село {total:.1f} км при плане "
                               f"{plan:.1f} км — остаток отнесён выше ПкСС")
            else:
                corridor, far_id, exact = tail_of(full, head["id"], plan)
                res["status"] = "tail" if exact else "tail_inexact"
                if not exact:
                    res["note"] = (f"хвост в {plan:.1f} км не ложится на узел: "
                                   f"ближайшая граница {path_km(corridor):.1f} км")

    # --- 3. Случай B: продлить коридор вверх по сети на излишек.
    if res["upstreamKm"] > TOL and far_id:
        first_edge_id = corridor[0]["id"] if corridor else None
        up, why = extend_upstream(graph, far_id, first_edge_id,
                                  res["upstreamKm"], guide)
        if up is None:
            res["status"] = "upstream_ambiguous"
            res["note"] = (f"{res['upstreamKm']:.1f} км выше «"
                           f"{graph.nodes[far_id].get('name')}» разложить нельзя — {why}")
        else:
            far_id = _far_of(graph, up, far_id)
            corridor = up + corridor

    # --- 3б. Голова недостижима — идём ОТ ОБРАТНОГО: от самого села вверх по
    # длинам участков свода. Часто последний отрезок в графе уже нарисован и
    # совпадает по длине (Сарыбулак: единственное ребро 15.5 км = весь план),
    # а не хватает только куска выше — тогда участок определён однозначно и
    # прятать его от карты незачем. Если ход упирается в развилку или обрыв,
    # статус остаётся прежним и строка идёт в отчёт.
    if res["status"] in ("no_route", "head_not_found"):
        back, why = extend_upstream(graph, snp["id"], None, plan,
                                    upstream_guide(rec["segs"], 0.0))
        if back:
            corridor, far_id = back, _far_of(graph, back, snp["id"])
            res["status"] = "from_np"
            res["note"] = (f"голова недоступна ({res['note']}); маршрут собран "
                           f"от села вверх по участкам свода")
            res["upstreamKm"] = 0.0
        else:
            res["note"] += f"; ход от села вверх тоже не вышел — {why}"
            return res

    res["corridorKm"] = round(path_km(corridor), 3)
    # Коридор в разы длиннее плана — это не участок, а промах якоря: метка ушла
    # к далёкому тёзке (Кызылорда, с.Билибай: план 2 км, а путь 736 км через 28
    # рёбер). Красить по такому нельзя — факт разъедется по всей области.
    if res["corridorKm"] > max(30.0, 8 * plan):
        res["status"] = "implausible_corridor"
        res["note"] = (f"коридор {res['corridorKm']:.1f} км при плане {plan:.1f} км "
                       f"({res['corridorKm'] / plan:.0f}× ) — якорь «"
                       f"{res['anchorLabel'] or rec['head']}» опознан неверно")
        res["edgeIds"] = []
        return res
    res["diffKm"] = round(res["corridorKm"] - plan, 3)
    res["edgeIds"] = [e["id"] for e in corridor]
    res["_corridor"] = corridor
    res["_farId"] = far_id

    # --- 4. Сверка именованных меток: лежат ли они на коридоре в нужном месте.
    res["segChecks"] = check_segments(graph, rec, corridor, far_id, snp)
    return res


def _far_of(graph: Graph, up: list[dict], anchor_id: str) -> str:
    """Дальний конец восходящего куска: идём от якоря назад по up."""
    cur = anchor_id
    for e in reversed(up):
        cur = other_end(e, cur)
    return cur


def check_segments(graph: Graph, rec: dict, corridor: list[dict],
                   far_id: str, snp: dict) -> list[dict]:
    """Для каждой ИМЕНОВАННОЙ метки: есть ли такой узел на коридоре и совпадает
    ли расстояние от неё до села с суммой отрезков свода. Это и есть проверка
    «правильность связей + плановая протяжённость» по отдельным участкам."""
    seq = node_seq(corridor, far_id)
    on_route = {nid: i for i, nid in enumerate(seq)}
    # расстояние от каждого узла коридора до села (идём с конца)
    to_snp, acc = {}, 0.0
    cur = snp["id"]
    to_snp[cur] = 0.0
    for e in reversed(corridor):
        acc += e.get("lengthKm") or 0.0
        cur = other_end(e, cur)
        to_snp[cur] = round(acc, 3)

    checks = []
    for j, s in enumerate(rec["segs"]):
        label = s["label"]
        if not label or is_placeholder(label):
            continue
        node = resolve_label(graph, label, rec["district"], near_id=snp["id"])
        expect = round(sum(x["km"] for x in rec["segs"][j:]), 3)
        item = {"label": label, "expectKm": expect}
        if node is None:
            item["state"] = "не опознана в графе"
        elif node["id"] not in on_route:
            item["state"] = "не лежит на коридоре"
            item["nodeName"] = node.get("name")
        else:
            got = to_snp.get(node["id"])
            item["graphKm"] = got
            item["nodeName"] = node.get("name")
            item["state"] = "ок" if got is not None and abs(got - expect) <= TOL \
                else "расстояние расходится"
        checks.append(item)
    return checks


# --------------------------------------------------------------- сборка ---
def build(graph: Graph, slug: str, title: str, recs: list[dict]) -> dict:
    bound = [bind(graph, r, recs) for r in recs]
    edge_by_id = {e["id"]: e for e in graph.edges}
    out_edges: dict[str, dict] = {}

    # Баланс км: каждый факт из свода обязан найтись в одной из четырёх
    # корзин, иначе цифры на карте молча разойдутся со сводом (см. отчёт).
    lost = {"unbound": {"tubeKm": 0.0, "fiberKm": 0.0},   # село не привязано
            "clipped": {"tubeKm": 0.0, "fiberKm": 0.0},   # факт > плана
            "spill": {"tubeKm": 0.0, "fiberKm": 0.0},     # коридор короче факта
            "edgecap": {"tubeKm": 0.0, "fiberKm": 0.0}}   # ребро уже заполнено

    PAINTABLE = {"chain", "chain_len_mismatch", "tail", "tail_inexact",
                 "tail_short", "from_np", "donor_segment", "head_to_np"}
    for si, b in enumerate(bound):
        corridor = b.pop("_corridor", None)
        far_id = b.pop("_farId", None)
        if not corridor or b["status"] not in PAINTABLE:
            lost["unbound"]["tubeKm"] += b["tubeKm"]
            lost["unbound"]["fiberKm"] += b["fiberKm"]
            continue
        # Факт обрезаем по плану: свод местами показывает больше, чем сам же
        # запланировал (сдают участок, не разнеся его по строкам). Излишек не
        # выдумываем, куда положить, — он уходит в отчёт.
        # Потолок — план строки, а у строк без плана (donor_segment) длина
        # участка, взятого из чужой строки.
        plan = b.get("capKm") or b["planKm"]
        tube = min(b["tubeKm"], plan)
        fiber = min(b["fiberKm"], plan)
        b["clippedTubeKm"] = round(max(0.0, b["tubeKm"] - tube), 3)
        b["clippedFiberKm"] = round(max(0.0, b["fiberKm"] - fiber), 3)
        lost["clipped"]["tubeKm"] += b["clippedTubeKm"]
        lost["clipped"]["fiberKm"] += b["clippedFiberKm"]

        for metric, km in (("tubeKm", tube), ("fiberKm", fiber)):
            if km <= 0:
                continue
            parts, left = fill_along(corridor, far_id, km)
            lost["spill"][metric] += left
            for eid, slot in parts.items():
                cell = out_edges.setdefault(eid, {
                    "tubeKm": 0.0, "fiberKm": 0.0,
                    "fillFrom": slot["fillFrom"], "settlements": [], "_by": {},
                })
                # Ключ — НОМЕР строки в bound, а не название: тёзок в области
                # хватает («Ақтөбе», «Еңбек»), и по имени вклады бы склеились.
                # Группируем ещё и по концу, ОТ которого идёт заливка, — см.
                # свод по рёбрам ниже.
                by = cell["_by"].setdefault(metric, {}).setdefault(slot["fillFrom"], {})
                by[si] = by.get(si, 0.0) + slot["km"]
                if b["name"] not in cell["settlements"]:
                    cell["settlements"].append(b["name"])

    # Свод по рёбрам. Магистраль делят несколько сёл, и КАЖДОЕ пишет свой факт
    # по всему своему маршруту — то есть один и тот же физический отрезок
    # попадает в несколько строк. Складывать их нельзя: у Копы 33.2 км и у
    # Еспе 23.1 км — это одна и та же трасса Таргап→Копа, а не 56 км. Поэтому
    # среди сёл, кладущих ребро С ОДНОГО конца, берём МАКСИМУМ (построено
    # столько, сколько прошёл самый дальний), а вклады с разных концов
    # складываем — их работы идут навстречу. Итог всё равно режем по длине.
    painted = [{"tubeKm": 0.0, "fiberKm": 0.0} for _ in bound]
    for eid, cell in out_edges.items():
        length = edge_by_id[eid].get("lengthKm")
        for metric in ("tubeKm", "fiberKm"):
            groups = cell["_by"].get(metric, {})
            raw = sum(sum(g.values()) for g in groups.values())
            total = sum(max(g.values()) for g in groups.values() if g)
            capped = min(total, length) if length else total
            lost["edgecap"][metric] += raw - capped
            cell[metric] = round(capped, 3)
            # Обрезку и «победу максимума» возвращаем сёлам, чтобы сумма по
            # сёлам совпадала с суммой по рёбрам (на ней стоит плитка KPI).
            scale = (capped / total) if total > 0 else 0.0
            for g in groups.values():
                if not g:
                    continue
                top = max(g.values())
                for si, km in g.items():
                    painted[si][metric] += (km if km >= top else 0.0) * scale
        del cell["_by"]

    for b, got in zip(bound, painted):
        b["paintedTubeKm"] = round(got["tubeKm"], 3)
        b["paintedFiberKm"] = round(got["fiberKm"], 3)

    # Охват графа считаем по ОБЪЕДИНЕНИЮ рёбер: наивная сумма коридоров кратно
    # переоценивает область, потому что магистраль делят десятки сёл ниже неё.
    touched = {eid for b in bound if b["status"] in PAINTABLE for eid in b["edgeIds"]}
    counts: dict[str, int] = {}
    for b in bound:
        counts[b["status"]] = counts.get(b["status"], 0) + 1

    return {
        "meta": {
            "slug": slug, "title": title,
            "source": "общий свод по смр СНП 2.0.xlsx",
            "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "metrics": {"tubeKm": "Магистр. Сеть (трубка) км",
                        "fiberKm": "Магистр. Сеть (ВОЛС) км"},
            "counts": counts,
        },
        "settlements": bound,
        "edges": out_edges,
        "totals": {
            "planKm": round(sum(b["planKm"] for b in bound), 1),
            "tubeKm": round(sum(b["tubeKm"] for b in bound), 1),
            "fiberKm": round(sum(b["fiberKm"] for b in bound), 1),
            "paintedTubeKm": round(sum(c["tubeKm"] for c in out_edges.values()), 1),
            "paintedFiberKm": round(sum(c["fiberKm"] for c in out_edges.values()), 1),
            "graphTouchedKm": round(
                sum(edge_by_id[eid].get("lengthKm") or 0.0 for eid in touched), 1),
            "graphPlannedKm": round(
                sum(e.get("lengthKm") or 0.0 for e in graph.edges
                    if e.get("type") == "planned"), 1),
            "lost": {k: {m: round(v, 1) for m, v in d.items()} for k, d in lost.items()},
        },
    }


# ---------------------------------------------------------------- отчёт ---
STATUS_RU = {
    "chain": "привязано по цепочке участков",
    "chain_len_mismatch": "цепочка есть, длина расходится",
    "tail": "хвост маршрута (в цепочке только П-N)",
    "tail_inexact": "хвост не ложится на узел",
    "tail_short": "путь короче плана, остаток выше ПкСС",
    "upstream_ambiguous": "излишек выше ПкСС — развилка, к разбору",
    "from_np": "собрано от села вверх (голова недоступна)",
    "donor_segment": "участок взят из строки соседнего НП",
    "head_to_np": "плана нет — взят отрезок ПкСС→село",
    "implausible_corridor": "коридор в разы длиннее плана — якорь неверен",
    "no_route": "нет пути в графе",
    "np_not_found": "села нет в графе",
    "head_not_found": "ПкСС не опознан",
    "skipped_no_plan": "пропущено: в своде нет протяжённости",
}
# Строка без плана — не проблема, а сознательный пропуск, поэтому в раздел
# «требуют разбора» она не идёт.
HARD = ("no_route", "np_not_found", "head_not_found", "upstream_ambiguous",
        "implausible_corridor")


def render_report(payload: dict) -> str:
    m, t = payload["meta"], payload["totals"]
    S = payload["settlements"]
    L: list[str] = []
    A = L.append

    A(f"# Сверка «{m['title']}» со сводом СНП 2.0\n")
    A(f"Источник: `{m['source']}` · собрано {m['generated']} · "
      f"`scripts/import_smr2.py`\n")
    A("Маршрут каждого села восстановлен по цепочке участков свода "
      "(метка → км → метка → … → НП), затем сопоставлен с графом. "
      "Факт — столбцы «Магистр. Сеть (трубка) км» и «Магистр. Сеть (ВОЛС) км», "
      "обрезанный по плану строки.\n")

    A("## Итого\n")
    A("| показатель | км |")
    A("|---|---:|")
    A(f"| План по своду | {t['planKm']} |")
    A(f"| Плановые рёбра графа (всего в области) | {t['graphPlannedKm']} |")
    A(f"| Рёбра, задетые коридорами (без повторов) | {t['graphTouchedKm']} |")
    A("")

    A("## Баланс факта\n")
    A("Каждый километр из свода обязан попасть в одну из строк, иначе карта "
      "молча разойдётся со сводом.\n")
    lo = t["lost"]
    A("| куда ушёл факт | трубка, км | ВОЛС, км |")
    A("|---|---:|---:|")
    A(f"| Всего в своде | {t['tubeKm']} | {t['fiberKm']} |")
    A(f"| — легло на рёбра карты | {t['paintedTubeKm']} | {t['paintedFiberKm']} |")
    A(f"| — село не привязано (раздел 1) | {lo['unbound']['tubeKm']} | "
      f"{lo['unbound']['fiberKm']} |")
    A(f"| — срезано, факт > плана (раздел 4) | {lo['clipped']['tubeKm']} | "
      f"{lo['clipped']['fiberKm']} |")
    A(f"| — коридор в графе короче факта | {lo['spill']['tubeKm']} | "
      f"{lo['spill']['fiberKm']} |")
    A(f"| — тот же отрезок уже учтён другим селом | {lo['edgecap']['tubeKm']} | "
      f"{lo['edgecap']['fiberKm']} |")
    A("")

    A("## Как привязались сёла\n")
    A("| статус | сёл |")
    A("|---|---:|")
    for k, v in sorted(m["counts"].items(), key=lambda kv: -kv[1]):
        A(f"| {STATUS_RU.get(k, k)} | {v} |")
    A("")

    def table(rows, cols, headers, empty="— нет —"):
        if not rows:
            A(empty + "\n")
            return
        A("| " + " | ".join(headers) + " |")
        A("|" + "|".join("---" for _ in headers) + "|")
        for r in rows:
            A("| " + " | ".join(str(c(r)) for c in cols) + " |")
        A("")

    chain = lambda s: " → ".join(  # noqa: E731
        f"{x['label'] or '?'} ({x['km']})" for x in s["segs"]) or "—"

    # Отображается только факт, поэтому сёла без факта в расхождениях не так
    # важны: они ничего не красят и ничего не искажают.
    fact = [s for s in S if (s["tubeKm"] or 0) > 0 or (s["fiberKm"] or 0) > 0]
    bad_fact = [s for s in fact if s["status"] not in ("chain", "tail")]
    A("## Сёла, у которых есть факт\n")
    A(f"Факт есть у **{len(fact)}** сёл из {len(S)} — на карте видны только они. "
      f"Из них привязано точно **{len(fact) - len(bad_fact)}**.\n")
    table(bad_fact,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["head"],
           lambda s: s["planKm"], lambda s: s["tubeKm"], lambda s: s["fiberKm"],
           lambda s: STATUS_RU.get(s["status"], s["status"]), lambda s: s["note"]],
          ["стр", "НП", "ПкСС", "план", "трубка", "ВОЛС", "статус", "что не так"],
          "— все сёла с фактом привязаны точно —")

    A("## 1. Требуют ручного разбора\n")
    A("Ничего не закрашено на карте: алгоритм не смог однозначно определить "
      "участок.\n")
    hard = [s for s in S if s["status"] in HARD]
    table(hard,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["district"],
           lambda s: s["head"], lambda s: s["planKm"],
           lambda s: STATUS_RU.get(s["status"], s["status"]), lambda s: s["note"]],
          ["стр", "НП", "район", "ПкСС", "план", "статус", "что не так"])

    A("## 2. Расхождение длины: граф vs свод\n")
    A("Коридор построен и закрашен, но его длина в графе не сходится с планом "
      "свода — либо в чертеже другая протяжённость, либо связь идёт не туда.\n")
    lm = sorted([s for s in S if s["status"] == "chain_len_mismatch"],
                key=lambda s: -abs(s["diffKm"] or 0))
    table(lm,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["anchorLabel"],
           lambda s: s["planKm"], lambda s: s["corridorKm"],
           lambda s: f"{s['diffKm']:+.1f}", lambda s: chain(s)],
          ["стр", "НП", "якорь", "план", "граф", "Δ", "цепочка свода"])

    A("## 3. Именованные участки, не легшие на коридор\n")
    A("Метка участка опознана в графе, но её нет на найденном маршруте или "
      "расстояние до села другое — прямой признак неверной связи.\n")
    bad = []
    for s in S:
        for c in s["segChecks"]:
            if c["state"] != "ок":
                bad.append((s, c))
    table(bad,
          [lambda p: p[0]["row"], lambda p: p[0]["name"], lambda p: p[1]["label"],
           lambda p: p[1].get("nodeName", "—"), lambda p: p[1]["expectKm"],
           lambda p: p[1].get("graphKm", "—"), lambda p: p[1]["state"]],
          ["стр", "НП", "метка свода", "узел графа", "свод, км", "граф, км", "состояние"])

    A("## 4. Факт больше плана (обрезано)\n")
    A("По договорённости лишние км не раскладываются: ребро красится максимум "
      "на 100 %, а излишек — сюда.\n")
    ov = sorted([s for s in S if s["clippedTubeKm"] > 0.05 or s["clippedFiberKm"] > 0.05],
                key=lambda s: -max(s["clippedTubeKm"], s["clippedFiberKm"]))
    table(ov,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["planKm"],
           lambda s: s["tubeKm"], lambda s: s["fiberKm"],
           lambda s: f"{s['clippedTubeKm']} / {s['clippedFiberKm']}"],
          ["стр", "НП", "план", "трубка", "ВОЛС", "срезано трубка / ВОЛС"])

    A("## 5. Хвост не лёг на узел\n")
    A("В цепочке одни П-N, поэтому взят хвост маршрута длиной в план, но план "
      "режет ребро посередине — граница участка в графе не там, где в своде.\n")
    ti = [s for s in S if s["status"] == "tail_inexact"]
    table(ti,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["head"],
           lambda s: s["planKm"], lambda s: s["corridorKm"], lambda s: s["note"]],
          ["стр", "НП", "ПкСС", "план", "коридор", "подробности"])

    A("## 6. Работы выше ПкСС (случай B)\n")
    A("Цепочка свода начинается выше самого ПкСС — эти км разложены вверх по "
      "сети, пока путь был однозначен.\n")
    up = [s for s in S if s["upstreamKm"] > TOL and s["status"] != "upstream_ambiguous"]
    table(up,
          [lambda s: s["row"], lambda s: s["name"], lambda s: s["head"],
           lambda s: s["planKm"], lambda s: s["upstreamKm"], lambda s: chain(s)],
          ["стр", "НП", "ПкСС", "план", "выше ПкСС, км", "цепочка свода"])

    return "\n".join(L) + "\n"


# ----------------------------------------------------------------- вход ---
GUARD_PLAN_DRIFT = 0.25
GUARD_BIND_DROP = 0.15


def suspicious(path: Path, payload: dict) -> str:
    """Причина не перезаписывать файл области, либо пустая строка.

    Свод правят руками, и одна сдвинутая колонка способна обнулить привязку по
    всей области. Раз обновление автоматическое и никто не смотрит на него
    каждый день, сравниваем с прошлым разбором и отказываемся затирать
    осмысленные данные подозрительными."""
    if not path.exists():
        return ""
    try:
        old = json.loads(path.read_text(encoding="utf-8"))
    except Exception:               # noqa: BLE001
        return ""

    def bind_share(p: dict) -> float:
        rows = p.get("settlements") or []
        if not rows:
            return 0.0
        return sum(1 for s in rows if s.get("status") in ("chain", "tail")) / len(rows)

    old_plan = (old.get("totals") or {}).get("planKm") or 0.0
    new_plan = (payload.get("totals") or {}).get("planKm") or 0.0
    if old_plan > 0 and abs(new_plan - old_plan) / old_plan > GUARD_PLAN_DRIFT:
        return (f"план области изменился с {old_plan} на {new_plan} км "
                f"— похоже на сдвиг столбцов")
    drop = bind_share(old) - bind_share(payload)
    if drop > GUARD_BIND_DROP:
        return (f"привязка просела с {bind_share(old):.0%} до "
                f"{bind_share(payload):.0%} сёл")
    return ""


def fetch_sheet(url: str = SHEET_URL) -> Path:
    """Скачать живую таблицу как xlsx во временный файл.

    Экспорт открыт по ссылке, ключи не нужны. Google кладёт в выгрузку и
    посчитанные значения формул, поэтому итоговые столбцы читаются как числа."""
    import urllib.request

    dst = Path(tempfile.gettempdir()) / "snp2-svod.xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "graph-vis/smr2"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        data = resp.read()
    if not data.startswith(b"PK"):
        raise SystemExit("Google вернул не xlsx — проверьте доступ по ссылке")
    dst.write_bytes(data)
    print(f"  скачано из Google: {len(data) / 1e6:.1f} МБ → {dst}")
    return dst


def find_xlsx() -> Path:
    for p in (BASE.parent / "data").glob("*.xlsx"):
        if "свод" in p.name.lower() and "снп" in p.name.lower():
            return p
    raise SystemExit("не нашёл «общий свод по смр СНП 2.0.xlsx» в data/")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--region", default="zhambyl",
                    help="slug области или несколько через запятую; "
                         "'all' — все листы свода")
    ap.add_argument("--from-sheet", action="store_true",
                    help="взять свод прямо из Google-таблицы, а не из data/")
    ap.add_argument("--report-only", action="store_true",
                    help="не писать smr2/*.json, только отчёт в stdout")
    args = ap.parse_args()

    # Консоль Windows по умолчанию в cp1251 и давится на «→» в логе прогона.
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")

    import openpyxl

    xlsx = fetch_sheet() if args.from_sheet else find_xlsx()
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)

    # Области разбираем по одной, но отчёт общий: перечислите уже разобранные
    # через запятую, иначе SMR2_RECONCILE.md потеряет их разделы.
    wanted = None if args.region == "all" else {
        s.strip() for s in args.region.split(",") if s.strip()}
    by_slug: dict[str, list[dict]] = {}
    titles: dict[str, str] = {}
    failed: dict[str, str] = {}

    for sheet, slug in SHEET_TO_SLUG.items():
        if sheet not in wb.sheetnames:
            continue
        if wanted and slug not in wanted:
            continue
        try:
            recs, control = parse_sheet(wb[sheet])
        except Exception as exc:        # noqa: BLE001
            failed[slug] = f"лист «{sheet}»: {exc}"
            print(f"  ! лист {sheet:<30} НЕ РАЗОБРАН: {exc}")
            continue
        by_slug.setdefault(slug, []).extend(recs)
        titles.setdefault(slug, sheet)
        print(f"  лист {sheet:<32} → {slug:<18} строк {len(recs)}")
        # Сверка с итоговой строкой самого листа. Расходится — либо наш разбор
        # съехал по столбцам, либо в книге битая формула: в «Алматинской» итог
        # трубки считает SUM(WQ3:WQ81) при данных до 82-й строки и теряет
        # последнее село. Молчать про такое нельзя — цифры пойдут в отчёт.
        for key, ru, mine in (
            ("np", "НП", float(len(recs))),
            ("planKm", "план", sum(r["planKm"] for r in recs)),
            ("tubeKm", "трубка", sum(r["tubeKm"] for r in recs)),
            ("fiberKm", "ВОЛС", sum(r["fiberKm"] for r in recs)),
        ):
            want = control.get(key)
            if want is None or abs(want - mine) <= 0.05:
                continue
            print(f"      ! {ru}: разбор {mine:.3f}, итог листа (стр "
                  f"{control.get('row')}) {want:.3f}, Δ={mine - want:+.3f}")

    if not by_slug:
        raise SystemExit(f"в своде нет листов для области «{args.region}»")

    if not args.report_only:
        OUT_DIR.mkdir(exist_ok=True)

    reports = []
    for slug, recs in sorted(by_slug.items()):
        gpath = REGIONS_DIR / f"{slug}.json"
        if not gpath.exists():
            print(f"  ! графа {slug}.json нет, пропускаю")
            continue
        graph = Graph(gpath)
        title = graph.meta.get("registry", {}).get("title") or titles[slug]
        payload = build(graph, slug, title, recs)
        t, c = payload["totals"], payload["meta"]["counts"]
        ok = c.get("chain", 0) + c.get("tail", 0)
        print(f"  {slug:<18} сёл {len(recs):>4} · привязано точно {ok:>4} · "
              f"план {t['planKm']:>8} км · трубка {t['tubeKm']:>7} км · "
              f"ВОЛС {t['fiberKm']:>7} км")
        if slug in failed:
            print(f"      ! {slug}: часть листов не разобрана, файл не трогаю")
        elif not args.report_only:
            why = suspicious(OUT_DIR / f"{slug}.json", payload)
            if why:
                failed[slug] = why
                print(f"      ! {slug}: {why} — файл НЕ перезаписан")
            else:
                (OUT_DIR / f"{slug}.json").write_text(
                    json.dumps(payload, ensure_ascii=False, indent=1),
                    encoding="utf-8")
        reports.append(render_report(payload))

    text = "\n\n---\n\n".join(reports)
    if args.report_only:
        sys.stdout.write(text)
    else:
        REPORT.write_text(text, encoding="utf-8")
        print(f"\nзаписано: smr2/*.json и {REPORT.name}")

    if failed:
        print("\nОБЛАСТИ БЕЗ ОБНОВЛЕНИЯ "
              "(на карте осталась прежняя версия):")
        for slug, why in sorted(failed.items()):
            print(f"  · {slug}: {why}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
