"""Импорт свода СМР из «Отчет СОИ*.xlsx» → smr/<slug>.json + отчёт о сверке.

ЗАЧЕМ. Вьюер до сих пор знал только два числа: сколько оцифровано в графе и
сколько сдано по отчётам работников в Telegram-боте. Утверждённый свод СМР
(план в км по каждому селу) и ФАКТ подрядчиков (магистральная трубка, задувка
ВОЛС, сданные СНП) живут в отдельной книге «Отчет СОИ», лист «СМР», и на карту
до сих пор не попадали. Этот скрипт кладёт их рядом с графом.

СТРУКТУРА ЛИСТА. Три уровня в одной колонке «Наименование области»:
область → район → село. Номер (`№п/п`) для определения уровня НЕПРИГОДЕН:
у большинства строк сёл он пуст, а часть номеров Excel превратил в даты
(«10.1» → 2026-01-10). Надёжный признак агрегата — заполненная колонка
«Количество СНП (шт)»: она есть у области и у района и пуста у села. Область
от района отличаем по целому номеру (1..17) — они как раз уцелели.

СОПОСТАВЛЕНИЕ С ГРАФОМ. Село свода → СНП-узел графа тем же нечётким матчером,
что и привязка отчётов (matcher.py: каз↔рус, «с.»/«ст.» отбрасываются), с тем
же порогом BIND_MIN=75 и подсказкой по району. Ниже порога — НЕ привязываем:
строка уходит в отчёт о сверке, а не ложится молча на чужое село (тот же
принцип, что и в progress_core: лучше честный «не найдено», чем километры
не в том районе).

Одна строка свода = одно село, но одна строка РАЙОНА свода может покрывать
несколько районов графа («Отырарский район, Шардаринский район») — район
используется только как подсказка матчеру, поэтому это не мешает.

Использование:
    python scripts/import_smr.py                       # ищет data/Отчет СОИ*.xlsx
    python scripts/import_smr.py path/to/Отчет.xlsx
    python scripts/import_smr.py --report-only         # только сверка, без записи
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

HERE = Path(__file__).resolve().parent
BASE = HERE.parent
sys.path.insert(0, str(BASE))

from matcher import district_stem, districts_compatible, match_score, normalize_name  # noqa: E402
from progress_core import Graph, PLANNED_TYPES, attribute_edges_to_snp  # noqa: E402

REGIONS_DIR = BASE / "regions"
SMR_DIR = BASE / "smr"
REPO_ROOT = BASE.parent
SHEET = "СМР"

# Область свода → slug графа. Часть графов сведена по две области (так их
# рисовали на чертежах), поэтому отображение не один-к-одному.
REGION_TO_SLUG = {
    "Туркестанская область": "turkestan",
    "Жамбылская область": "zhambyl",
    "Область Ұлытау": "karaganda-ulytau",
    "Карагандинская область": "karaganda-ulytau",
    "Кызылординская область": "kyzylorda",
    "Атырауская область": "mangystau-atyrau",
    "Мангистауская область": "mangystau-atyrau",
    "Область Абай": "abai",
    "Северо-Казахстанская область": "sko",
    "Западно-Казахстанская область": "zko",
    "Восточно-Казахстанская область": "vko",
    "Акмолинская область": "akmola",
    "Алматинская область": "almaty-zhetysu",
    "Область Жетысу": "almaty-zhetysu",
    "Павлодарская область": "pavlodar",
    "Костанайская область": "kostanay",
    "Актюбинская область": "aktobe",
}

# Колонки листа (0-based). ПЛАН — «Протяженность ВОЛС (км)»; ФАКТ — накопительный
# итог слева от дневных колонок (дневные не читаем: нас интересует состояние на
# сегодня, а не разбивка по датам).
C_CONTRACTOR, C_NUM, C_NAME, C_DISTRICTS, C_SNP_PLAN, C_PLAN_KM = 0, 1, 2, 3, 4, 5
F_TUBE, F_FIBER, F_SNP, F_GUBO, F_B2C = 7, 8, 9, 10, 11
FIRST_DAILY_COL = 13  # с этой колонки идут дневные срезы — из них берём только дату

BIND_MIN = 75  # тот же порог привязки, что в progress_core

_INT_RE = re.compile(r"^\d+$")
_SETTLEMENT_PREFIX_RE = re.compile(r"^\s*(с|ст|п|пос|аул|село|рзд|разъезд|мкр)\s*\.", re.I)
_ADMIN_WORD_RE = re.compile(r"\b(район|ауданы|область)\b", re.I)


def _num(v):
    """Число из ячейки; пусто/мусор → None. В книге встречаются и числа,
    и строки с пробелами-разделителями тысяч, и запятая вместо точки."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def district_parts(raw: str | None) -> list[str]:
    """«Отырарский район, Шардаринский район» → обе части (одна строка свода
    нередко покрывает два района графа)."""
    return [p.strip() for p in (raw or "").split(",") if p.strip()]


def name_variants(raw: str) -> list[str]:
    """Варианты названия села из одной ячейки свода.

    В своде массово встречается «новое (старое)»: «Ғарышкер (с.Гагарин)»,
    «Қостақыр (с.60 лет Казахстан)». В графе может быть записан любой из двух,
    поэтому пробуем оба, а не только строку целиком."""
    s = _clean(raw).rstrip(",;").strip()
    out = [s]
    m = re.match(r"^(.*?)\s*\((.*)\)\s*$", s)
    if m:
        out += [x.strip() for x in (m.group(1), m.group(2)) if x.strip()]
    for v in list(out):
        if "/" in v:
            out += [p.strip() for p in v.split("/") if p.strip()]
    return [v for v in dict.fromkeys(out) if v]


def is_subheader(name: str, cur_district: str | None) -> bool:
    """Строка-подзаголовок внутри блока района («г.Балхаш», «Илийский»,
    «Жанибекский район»), а не село. Такие строки в своде появляются, когда
    один блок покрывает несколько районов. Село с названием как у района
    («с.Аксу» в Аксуском районе) под правило не попадает — у него есть
    префикс «с.»."""
    if _SETTLEMENT_PREFIX_RE.match(name):
        return False
    if _ADMIN_WORD_RE.search(name):
        return True
    stem = district_stem(name)
    return bool(stem) and any(stem == district_stem(p) for p in district_parts(cur_district))


# --------------------------------------------------------------------------
# 1. Разбор книги
# --------------------------------------------------------------------------
def parse_smr(xlsx: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"в книге нет листа «{SHEET}» (есть: {', '.join(wb.sheetnames)})")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        sys.exit(f"лист «{SHEET}» пуст")

    # Дата среза — последняя дата в шапке дневных колонок (строка 2 книги).
    header = rows[1]
    dates = [c for c in header[FIRST_DAILY_COL:] if isinstance(c, datetime)]
    as_of = max(dates).date().isoformat() if dates else None

    regions, districts, setts, skipped = [], [], [], []
    cur_region = cur_district = None
    for i, r in enumerate(rows[3:], start=4):
        def g(j):
            return r[j] if j < len(r) else None

        name = _clean(g(C_NAME))
        if not name or name.upper() == "ИТОГО":
            continue
        rec = {
            "row": i,
            "name": name,
            "planKm": _num(g(C_PLAN_KM)),
            "tubeKm": _num(g(F_TUBE)),
            "fiberKm": _num(g(F_FIBER)),
            "snpDone": _num(g(F_SNP)),
            "guboDone": _num(g(F_GUBO)),
            "b2cDone": _num(g(F_B2C)),
        }
        aggregate = g(C_SNP_PLAN) not in (None, "")
        num = _clean(g(C_NUM))
        if aggregate and _INT_RE.match(num):
            cur_region, cur_district = name, None
            rec.update(level="region", districtsPlan=_num(g(C_DISTRICTS)),
                       snpPlan=_num(g(C_SNP_PLAN)))
            regions.append(rec)
        elif aggregate:
            cur_district = name
            rec.update(level="district", region=cur_region,
                       snpPlan=_num(g(C_SNP_PLAN)))
            districts.append(rec)
        elif is_subheader(name, cur_district):
            skipped.append({"row": i, "name": name, "region": cur_region,
                            "district": cur_district,
                            "reason": "подзаголовок района внутри блока, не село"})
        else:
            rec.update(level="settlement", region=cur_region, district=cur_district,
                       contractor=_clean(g(C_CONTRACTOR)) or None)
            setts.append(rec)
    return {"asOf": as_of, "regions": regions, "districts": districts,
            "settlements": setts, "skipped": skipped}


# --------------------------------------------------------------------------
# 2. Привязка сёл свода к СНП-узлам графа
# --------------------------------------------------------------------------
class GraphIndex:
    def __init__(self, path: Path):
        data = json.loads(path.read_text(encoding="utf-8"))
        self.slug = data["meta"]["registry"]["slug"]
        self.title = data["meta"]["registry"]["title"]
        self.snps = [n for n in data["nodes"] if n.get("kind") == "snp"]
        self.index = [(normalize_name(n.get("name")), normalize_name(n.get("district")), n)
                      for n in self.snps if n.get("name")]
        self.path = path
        self._graph: Graph | None = None
        self._branch_km: dict[str, float] | None = None

    def graph(self) -> Graph:
        if self._graph is None:
            self._graph = Graph(self.path)
        return self._graph

    def branch_km(self) -> dict[str, float]:
        """{id СНП: длина его собственной ветки плановых линий, км} — то, чем
        граф отвечает на строку свода «Протяженность ВОЛС (км)» по этому селу.
        Делёж рёбер между сёлами — тот же, что рисует вьюер (compute_smr)."""
        if self._branch_km is None:
            graph = self.graph()
            by_id = {e["id"]: e for e in graph.edges}
            km: dict[str, float] = {}
            for eid, snp_id in attribute_edges_to_snp(graph).items():
                km[snp_id] = km.get(snp_id, 0.0) + (by_id[eid].get("lengthKm") or 0.0)
            self._branch_km = km
        return self._branch_km

    def planned_km(self) -> float:
        return sum(e.get("lengthKm") or 0.0
                   for e in self.graph().edges if e.get("type") in PLANNED_TYPES)

    def best(self, raw_name: str, raw_district: str | None) -> tuple[dict | None, int]:
        """Лучший СНП-узел под название села свода — та же шкала очков, что у
        Graph.find_node_scored: совпадение района даёт бонус, явно чужой район —
        штраф (иначе тёзка из соседнего района перетянет километры на себя)."""
        parts = district_parts(raw_district) or [None]
        best_node, best_sc = None, 0
        for variant in name_variants(raw_name):
            qn = normalize_name(variant)
            if len(qn) < 2:
                continue
            for part in parts:
                dn = normalize_name(part)
                for nn, ndn, node in self.index:
                    sc = match_score(qn, nn)
                    if sc <= 0:
                        continue
                    if dn and ndn:
                        sc += 8 if districts_compatible(dn, ndn) else -15
                    if sc > best_sc:
                        best_node, best_sc = node, sc
        return best_node, best_sc


def build(parsed: dict, graphs: dict[str, GraphIndex]) -> dict[str, dict]:
    """Свод → {slug: payload smr/<slug>.json}."""
    out: dict[str, dict] = {}
    for slug, gi in graphs.items():
        out[slug] = {
            "meta": {"slug": slug, "title": gi.title, "asOf": parsed["asOf"],
                     "regions": [], "totals": {}},
            "byDistrict": {},
            "settlements": [],
            "unmatched": [],
        }

    for reg in parsed["regions"]:
        slug = REGION_TO_SLUG.get(reg["name"])
        if slug and slug in out:
            out[slug]["meta"]["regions"].append({
                "name": reg["name"], "planKm": reg["planKm"], "tubeKm": reg["tubeKm"],
                "fiberKm": reg["fiberKm"], "snpDone": reg["snpDone"],
                "snpPlan": reg["snpPlan"], "districtsPlan": reg["districtsPlan"],
            })

    for d in parsed["districts"]:
        slug = REGION_TO_SLUG.get(d.get("region") or "")
        if slug and slug in out:
            out[slug]["byDistrict"][d["name"]] = {
                "planKm": d["planKm"], "tubeKm": d["tubeKm"], "fiberKm": d["fiberKm"],
                "snpDone": d["snpDone"], "snpPlan": d["snpPlan"],
            }

    for s in parsed["settlements"]:
        slug = REGION_TO_SLUG.get(s.get("region") or "")
        if not slug or slug not in out:
            continue
        node, score = graphs[slug].best(s["name"], s.get("district"))
        row = {k: s[k] for k in ("name", "planKm", "tubeKm", "fiberKm",
                                 "snpDone", "guboDone", "b2cDone")}
        row.update(district=s.get("district"), region=s.get("region"),
                   contractor=s.get("contractor"), row=s["row"])
        if node is not None and score >= BIND_MIN:
            row.update(nodeId=node["id"], nodeName=node.get("name"),
                       nodeDistrict=node.get("district"), score=score,
                       # Чем граф отвечает на плановые км свода — длина ветки
                       # этого села; расхождение и есть предмет сверки связей.
                       branchKm=round(graphs[slug].branch_km().get(node["id"], 0.0), 2))
            out[slug]["settlements"].append(row)
        else:
            row.update(score=score,
                       nearest=(node.get("name") if node else None),
                       nearestDistrict=(node.get("district") if node else None),
                       reason=("нет похожего СНП в графе" if node is None else
                               f"лучший кандидат «{node.get('name')}» "
                               f"({node.get('district')}) — счёт {score} < {BIND_MIN}"))
            out[slug]["unmatched"].append(row)

    for slug, payload in out.items():
        agg = {"planKm": 0.0, "tubeKm": 0.0, "fiberKm": 0.0,
               "snpDone": 0.0, "guboDone": 0.0, "b2cDone": 0.0}
        for r in payload["meta"]["regions"]:
            for k in agg:
                agg[k] += r.get(k) or 0
        payload["meta"]["totals"] = {k: round(v, 3) for k, v in agg.items()}
        payload["meta"]["counts"] = {
            "settlementsInSmr": len(payload["settlements"]) + len(payload["unmatched"]),
            "matched": len(payload["settlements"]),
            "unmatched": len(payload["unmatched"]),
            "snpInGraph": len(graphs[slug].snps),
        }
    return out


# --------------------------------------------------------------------------
# 3. Отчёт о сверке
# --------------------------------------------------------------------------
def render_report(parsed: dict, built: dict[str, dict], graphs: dict[str, GraphIndex],
                  xlsx: Path) -> str:
    L = [
        "# Сверка свода СМР с графами",
        "",
        f"Источник: `{xlsx.name}`, лист «{SHEET}»"
        + (f", срез на {parsed['asOf']}" if parsed["asOf"] else ""),
        f"Сгенерировано: {datetime.now(timezone.utc).date().isoformat()} "
        f"(`python scripts/import_smr.py`)",
        "",
        "Файл собирается скриптом — правьте не его, а исходную книгу или граф.",
        "",
        "## Итог по областям",
        "",
        "| граф | сёл в своде | привязано | не привязано | СНП в графе | нет в своде |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for slug in sorted(built):
        p = built[slug]
        c = p["meta"]["counts"]
        covered = {r["nodeId"] for r in p["settlements"]}
        L.append(f"| {slug} | {c['settlementsInSmr']} | {c['matched']} | "
                 f"{c['unmatched']} | {c['snpInGraph']} | "
                 f"{c['snpInGraph'] - len(covered)} |")

    L += ["", "## Связи: плановые км свода против оцифрованных", "",
          "Свод даёт «Протяженность ВОЛС (км)» на каждое село; граф отвечает "
          "суммой длин ВЕТКИ этого села — плановых линий, которые ближе к нему, "
          "чем к любому другому селу (тот же делёж, что рисует вьюер). Совпадать "
          "до метра они не обязаны — свод считает трассу по местности, чертёж "
          "рисует сегменты, — но большой разрыв означает, что участок либо не "
          "оцифрован, либо оцифрован не туда.", "",
          "| граф | план свода, км | ветки графа, км | Δ | плановых линий графа, км |",
          "|---|---:|---:|---:|---:|"]
    for slug in sorted(built):
        rows = built[slug]["settlements"]
        smr_km = sum(r.get("planKm") or 0 for r in rows)
        br_km = sum(r.get("branchKm") or 0 for r in rows)
        diff = br_km - smr_km
        L.append(f"| {slug} | {smr_km:.1f} | {br_km:.1f} | {diff:+.1f} | "
                 f"{graphs[slug].planned_km():.1f} |")

    L += ["", "### Сёла с наибольшим расхождением (|Δ| ≥ 5 км)", "",
          "Первые 25 по каждой области — с них и стоит начинать сверку чертежа.", ""]
    for slug in sorted(built):
        rows = [r for r in built[slug]["settlements"]
                if r.get("planKm") is not None
                and abs((r.get("branchKm") or 0) - r["planKm"]) >= 5]
        if not rows:
            continue
        rows.sort(key=lambda r: -abs((r.get("branchKm") or 0) - r["planKm"]))
        L.append(f"#### {slug} — {len(rows)}")
        L.append("")
        L.append("| село (свод) | узел графа | район графа | план свода, км | ветка графа, км | Δ |")
        L.append("|---|---|---|---:|---:|---:|")
        for r in rows[:25]:
            d = (r.get("branchKm") or 0) - r["planKm"]
            L.append(f"| {r['name']} | {r.get('nodeName') or r['nodeId']} | "
                     f"{r.get('nodeDistrict') or '—'} | {r['planKm']} | "
                     f"{r.get('branchKm')} | {d:+.1f} |")
        L.append("")

    L += ["", "## Сёла свода, не найденные в графе", "",
          "Строка свода есть, а СНП-узла под неё в графе нет (или совпадение "
          "слишком слабое). Либо село не оцифровано на чертеже, либо названия "
          "разошлись — тогда достаточно переименовать узел в графе.", ""]
    for slug in sorted(built):
        rows = built[slug]["unmatched"]
        if not rows:
            continue
        L.append(f"### {slug} — {len(rows)}")
        L.append("")
        L.append("| район свода | название в своде | план, км | ближайший узел графа | счёт |")
        L.append("|---|---|---:|---|---:|")
        for r in sorted(rows, key=lambda x: (x.get("district") or "", x["name"])):
            near = r.get("nearest") or "—"
            if r.get("nearestDistrict"):
                near += f" ({r['nearestDistrict']})"
            L.append(f"| {r.get('district') or '—'} | {r['name']} | "
                     f"{r.get('planKm') if r.get('planKm') is not None else '—'} | "
                     f"{near} | {r.get('score') or 0} |")
        L.append("")

    L += ["## СНП графа, которых нет в своде", "",
          "Узел на чертеже есть, а строки в своде СМР под него нет. Это либо "
          "село вне проекта подключения, либо пропуск в своде.", ""]
    for slug in sorted(built):
        gi = graphs[slug]
        covered = {r["nodeId"] for r in built[slug]["settlements"]}
        rest = [n for n in gi.snps if n["id"] not in covered]
        if not rest:
            continue
        L.append(f"### {slug} — {len(rest)}")
        L.append("")
        L.append("| id | название | район | тип |")
        L.append("|---|---|---|---|")
        for n in sorted(rest, key=lambda n: ((n.get("district") or ""), (n.get("name") or ""))):
            L.append(f"| {n['id']} | {n.get('name') or '—'} | "
                     f"{n.get('district') or '—'} | {n.get('subtype') or '—'} |")
        L.append("")

    if parsed["skipped"]:
        L += ["## Строки свода, пропущенные как подзаголовки", "",
              "Внутри блока района свод иногда перечисляет соседние районы "
              "отдельными строками — это не сёла.", "",
              "| строка | название | блок района |", "|---:|---|---|"]
        for s in parsed["skipped"]:
            L.append(f"| {s['row']} | {s['name']} | {s.get('district') or '—'} |")
        L.append("")
    return "\n".join(L)


# --------------------------------------------------------------------------
def find_xlsx() -> Path:
    candidates = sorted((REPO_ROOT / "data").glob("Отчет СОИ*.xlsx"))
    if not candidates:
        sys.exit("не нашёл data/Отчет СОИ*.xlsx — укажите путь аргументом")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", nargs="?", help="книга «Отчет СОИ»; по умолчанию — свежая из data/")
    ap.add_argument("--report-only", action="store_true",
                    help="только напечатать сверку, ничего не записывать")
    args = ap.parse_args()

    xlsx = Path(args.xlsx) if args.xlsx else find_xlsx()
    if not xlsx.exists():
        sys.exit(f"нет файла {xlsx}")

    graphs = {}
    for p in sorted(REGIONS_DIR.glob("*.json")):
        gi = GraphIndex(p)
        graphs[gi.slug] = gi

    parsed = parse_smr(xlsx)
    unknown = {r["name"] for r in parsed["regions"] if r["name"] not in REGION_TO_SLUG}
    if unknown:
        print(f"⚠ области свода без графа (пропущены): {', '.join(sorted(unknown))}",
              file=sys.stderr)

    built = build(parsed, graphs)
    report = render_report(parsed, built, graphs, xlsx)

    print(f"книга: {xlsx.name}, срез на {parsed['asOf']}")
    print(f"свод: областей {len(parsed['regions'])}, районов {len(parsed['districts'])}, "
          f"сёл {len(parsed['settlements'])}, подзаголовков пропущено "
          f"{len(parsed['skipped'])}")
    for slug in sorted(built):
        c = built[slug]["meta"]["counts"]
        t = built[slug]["meta"]["totals"]
        print(f"  {slug:<18} сёл {c['settlementsInSmr']:>4} → привязано {c['matched']:>4}, "
              f"не привязано {c['unmatched']:>3} · план {t['planKm']:>8.1f} км · "
              f"трубка {t['tubeKm']:>7.1f} · ВОЛС {t['fiberKm']:>6.1f}")

    if args.report_only:
        return

    SMR_DIR.mkdir(exist_ok=True)
    for slug, payload in built.items():
        (SMR_DIR / f"{slug}.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    (BASE / "SMR_RECONCILE.md").write_text(report, encoding="utf-8")
    print(f"\nзаписано: smr/*.json ({len(built)} файлов) и SMR_RECONCILE.md")


if __name__ == "__main__":
    main()
